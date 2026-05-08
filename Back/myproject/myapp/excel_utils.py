import os
from openpyxl import load_workbook, Workbook
from django.conf import settings
from django.db.models import Sum
from .models import User_WorkDay, Expense
from datetime import date
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
import calendar


# ----------------------
# 공통 워크북 헬퍼
# ----------------------

def _get_workbook(template_file, workplace_name=None, default_template="default.xlsx"):
    """템플릿을 로드하거나 새로 생성합니다."""
    if template_file:
        try:
            return load_workbook(template_file)
        except Exception as e:
            print(f"Failed to load provided template: {e}")
    
    # 로컬 템플릿 로드 시도
    base_dir = settings.BASE_DIR
    template_dir = os.path.join(base_dir, 'workload')
    
    if workplace_name:
        template_path = os.path.join(template_dir, f"{workplace_name}.xlsx")
        if os.path.exists(template_path):
            return load_workbook(template_path)
            
    default_path = os.path.join(template_dir, default_template)
    if os.path.exists(default_path):
        return load_workbook(default_path)
        
    return Workbook()


# ----------------------
# 공통 워크시트 헬퍼
# ----------------------

def safe_set(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value


# ----------------------
# 공통 행 복사 헬퍼
# ----------------------

def copy_cell(src, dst):
    """셀 값, 스타일, 수식, 서식 복사"""
    if src.has_style:
        dst._style = copy(src._style)

    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)

    if src.data_type == "f":
        dst.value = Translator(
            src.value,
            origin=src.coordinate
        ).translate_formula(dst.coordinate)
    else:
        dst.value = src.value


def copy_row_block(ws, source_start, source_end, target_start, start_col=1, end_col=60):
    """행 블록 복사: 글, 스타일, 수식 포함"""
    row_gap = target_start - source_start

    for source_row in range(source_start, source_end + 1):
        target_row = source_row + row_gap

        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height

        for col in range(start_col, end_col + 1):
            src = ws.cell(source_row, col)
            dst = ws.cell(target_row, col)

            if isinstance(dst, MergedCell):
                continue

            if isinstance(src, MergedCell):
                continue

            copy_cell(src, dst)


# ----------------------
# GoogleDriveWorkplaceExcelExportAPIView
# ----------------------

def hour_value(minutes):
    if not minutes:
        return None
    value = minutes / 60
    return int(value) if value.is_integer() else value


WORK_TYPE_ROW = {
    "주간": 0,
    "잔업": 1,
    "중식연장": 2,
    "특근": 3,
    "철야": 4,
    "철야연장": 5,
    "조기출근": 6,
}


def copy_merged_cells(ws, source_start, source_end, target_start):
    """기준 블록 안의 병합셀을 새 블록에도 동일하게 생성"""
    row_gap = target_start - source_start

    source_ranges = list(ws.merged_cells.ranges)

    for merged_range in source_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # 기준 블록 안에 있는 병합셀만 복사
        if source_start <= min_row and max_row <= source_end:
            new_min_row = min_row + row_gap
            new_max_row = max_row + row_gap

            new_range = (
                f"{get_column_letter(min_col)}{new_min_row}:"
                f"{get_column_letter(max_col)}{new_max_row}"
            )

            if new_range not in ws.merged_cells:
                ws.merge_cells(new_range)


def ensure_worker_blocks(ws, worker_count):
    """
    템플릿에 2명까지 있음:
    1번 사람: 5~13행
    2번 사람: 14~22행

    3명 이상이면 23행 위에 9줄씩 추가
    """

    block_size = 9

    # 복사 기준: 2번째 사람 블록
    source_start = 14
    source_end = 22

    # 합계 또는 다음 영역 시작 행
    insert_at_row = 23

    extra_count = max(0, worker_count - 2)

    if extra_count <= 0:
        return

    # 필요한 줄 한 번에 추가
    ws.insert_rows(insert_at_row, amount=extra_count * block_size)

    # 추가된 각 사람 블록에 2번째 사람 블록 복사
    for i in range(extra_count):
        target_start = insert_at_row + (i * block_size)

        copy_row_block(
            ws,
            source_start=source_start,
            source_end=source_end,
            target_start=target_start,
            start_col=1,
            end_col=60,
        )

        copy_merged_cells(
            ws,
            source_start=source_start,
            source_end=source_end,
            target_start=target_start,
        )


def generate_workplace_excel(work_place, year, month, template_file=None):
    wb = _get_workbook(template_file, workplace_name=work_place)
    ws = wb.active

    work_days = User_WorkDay.objects.filter(
        work_place=work_place,
        work_date__year=year,
        work_date__month=month,
        is_approved=True
    ).prefetch_related("details")

    user_work_map = {}

    for wd in work_days:
        name = wd.user_name
        day = wd.work_date.day

        if name not in user_work_map:
            user_work_map[name] = {}

        if day not in user_work_map[name]:
            user_work_map[name][day] = {}

        for detail in wd.details.all():
            work_type = detail.work_type

            if work_type not in WORK_TYPE_ROW:
                continue

            row_offset = WORK_TYPE_ROW[work_type]

            if work_type in ["주간", "특근", "철야"]:
                value = 1
            else:
                value = hour_value(detail.minutes)

            if value is None:
                continue

            user_work_map[name][day][row_offset] = (
                user_work_map[name][day].get(row_offset, 0) + value
            )

    sorted_names = sorted(user_work_map.keys())

    # 핵심: 사람 수만큼 블록 확보
    ensure_worker_blocks(ws, len(sorted_names))

    _, last_day = calendar.monthrange(year, month)
    days_kr = ["월", "화", "수", "목", "금", "토", "일"]

    title_row = 2
    title_col = 4        # D2

    header_row_day = 3   # P3
    header_row_date = 4  # P4

    data_start_row = 5
    name_col = 4         # D열
    date_start_col = 16  # P열

    block_size = 9

    safe_set(ws, title_row, title_col, f"{year}년 {month}월 {work_place} 근무 현황")

    # 요일 / 날짜
    for day in range(1, last_day + 1):
        col = date_start_col + day - 1
        current_date = date(year, month, day)

        safe_set(ws, header_row_day, col, days_kr[current_date.weekday()])
        safe_set(ws, header_row_date, col, current_date.strftime("%m/%d"))

    # 사람별 데이터 입력
    for idx, name in enumerate(sorted_names):
        current_row = data_start_row + (idx * block_size)

        safe_set(ws, current_row, name_col, name)

        # 날짜 영역 초기화
        for row_offset in range(block_size):
            row = current_row + row_offset

            for day in range(1, last_day + 1):
                col = date_start_col + day - 1
                safe_set(ws, row, col, None)

        # 실제 근무 데이터 입력
        day_map = user_work_map[name]

        for day, type_map in day_map.items():
            col = date_start_col + day - 1

            for row_offset, value in type_map.items():
                safe_set(
                    ws,
                    current_row + row_offset,
                    col,
                    value
                )

    return wb


# ----------------------
# GoogleDriveSalaryExcelExportAPIView
# ----------------------

def find_salary_total_row(ws, data_start_row=5):
    # 템플릿에서 "총계"가 있는 행을 찾아 직원 데이터가 끝나는 위치를 확인
    for row in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, str) and "총" in value:
            return row
    return ws.max_row + 1


def ensure_salary_rows(ws, employee_count, data_start_row=5, total_row=None):
    if total_row is None:
        total_row = find_salary_total_row(ws, data_start_row=data_start_row)

    # 현재 템플릿에 들어갈 수 있는 직원 수보다 실제 직원 수가 많으면 총계행 위에 행 추가
    capacity = max(total_row - data_start_row, 0)
    extra_count = employee_count - capacity

    if extra_count <= 0:
        return total_row

    source_row = max(total_row - 1, data_start_row)

    # 총계행이 병합되어 있으면 행 삽입 전에 풀고, 삽입 후 새 총계행 위치에 다시 병합
    total_row_merges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == total_row and merged_range.max_row == total_row:
            total_row_merges.append(
                (
                    merged_range.min_col,
                    merged_range.max_col,
                )
            )
            ws.unmerge_cells(str(merged_range))

    ws.insert_rows(total_row, amount=extra_count)

    # 새로 추가한 직원 행은 기존 마지막 직원 행의 스타일과 수식을 복사
    for idx in range(extra_count):
        target_row = total_row + idx
        copy_row_block(
            ws,
            source_start=source_row,
            source_end=source_row,
            target_start=target_row,
            start_col=1,
            end_col=max(ws.max_column, 21),
        )

    shifted_total_row = total_row + extra_count
    for min_col, max_col in total_row_merges:
        ws.merge_cells(
            start_row=shifted_total_row,
            start_column=min_col,
            end_row=shifted_total_row,
            end_column=max_col,
        )

    return shifted_total_row


def generate_salary_excel(year, month, template_file=None):
    wb = _get_workbook(template_file, default_template="salary_template.xlsx")
    ws = wb.active

    # 승인된 근무일에 연결된 급여 지출(Expense)을 직원별로 합산
    salary_rows = list(
        Expense.objects
        .filter(
            work_day__isnull=False,
            work_day__is_approved=True,
            work_day__work_date__year=year,
            work_day__work_date__month=month,
        )
        .values(
            "work_day__user_uuid_id",
            "work_day__user_uuid__user_name",
            "work_day__user_uuid__resident_number",
        )
        .annotate(total_amount=Sum("amount"))
        .order_by("work_day__user_uuid__user_name")
    )

    # 템플릿 기준: 5행부터 직원 데이터, "총계" 행 전까지 직원 목록 입력
    data_start_row = 5
    total_row = find_salary_total_row(ws, data_start_row=data_start_row)
    company_name = ws.cell(row=data_start_row, column=2).value
    total_row = ensure_salary_rows(
        ws,
        employee_count=len(salary_rows),
        data_start_row=data_start_row,
        total_row=total_row,
    )

    safe_set(ws, 1, 1, f"({year}년 {month}월) 급.상여지급대장")

    # 기존 템플릿에 들어 있던 샘플 직원 데이터 초기화
    for row in range(data_start_row, total_row):
        for col in range(1, max(ws.max_column, 21) + 1):
            safe_set(ws, row, col, None)

    # 직원별 이름, 주민등록번호, 해당 월 월급액 입력
    for idx, salary in enumerate(salary_rows, start=1):
        row = data_start_row + idx - 1
        amount = salary["total_amount"] or 0

        safe_set(ws, row, 1, idx)
        safe_set(ws, row, 2, company_name)
        safe_set(ws, row, 3, salary["work_day__user_uuid__user_name"])
        safe_set(ws, row, 4, salary["work_day__user_uuid__resident_number"])
        safe_set(ws, row, 7, amount)
        safe_set(ws, row, 8, amount)
        safe_set(ws, row, 9, 0)
        safe_set(ws, row, 10, 0)
        safe_set(ws, row, 11, f"=SUM(H{row}:J{row})")

    last_data_row = max(data_start_row, total_row - 1)

    # 총계행은 직원 데이터 범위에 맞춰 다시 계산되도록 수식 입력
    safe_set(ws, total_row, 1, "총   계")
    for col in range(7, max(ws.max_column, 21) + 1):
        col_letter = get_column_letter(col)
        safe_set(ws, total_row, col, f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})")

    return wb


# ----------------------
# 개인 급여/근무기록 엑셀 임시 로직
# ----------------------

def generate_user_pay_excel(user_uuid, year, month, template_file=None):
    """관리자용: 특정 유저의 월급 명세서 엑셀 생성 (뼈대)"""
    wb = _get_workbook(template_file, default_template="user_pay_template.xlsx")
    ws = wb.active
    # TODO: 유저 월급 데이터 채우기
    ws['B2'] = f"{year}년 {month}월 급여 명세서"
    return wb

def generate_user_record_excel(user_uuid, year, month, template_file=None):
    """사용자용: 자신의 근무 기록 엑셀 생성 (뼈대)"""
    wb = _get_workbook(template_file, default_template="user_record_template.xlsx")
    ws = wb.active
    # TODO: 개인 근무 기록 데이터 채우기
    ws['B2'] = f"{year}년 {month}월 개인 근무 기록"
    return wb
