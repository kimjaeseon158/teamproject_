import struct
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from django.conf import settings
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

SCHEDULE_SHEET_NAME = "근무표"


class WorkScheduleValidationError(Exception):
    pass


class WorkScheduleConversionError(Exception):
    pass


@dataclass(frozen=True)
class PreviewPage:
    content: bytes
    width: int
    height: int


@dataclass(frozen=True)
class PreparedWorkSchedule:
    original_content: bytes
    original_name: str
    original_size: int
    preview_pages: list[PreviewPage]


def _png_dimensions(content):
    if len(content) < 24 or content[:8] != b"\x89PNG\r\n\x1a\n":
        raise WorkScheduleConversionError("미리보기 이미지 형식이 올바르지 않습니다.")
    return struct.unpack(">II", content[16:24])


def _run_command(command, *, timeout):
    try:
        return subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except FileNotFoundError as exc:
        raise WorkScheduleConversionError(
            f"변환 프로그램을 찾을 수 없습니다: {command[0]}"
        ) from exc
    except subprocess.TimeoutExpired as exc:
        raise WorkScheduleConversionError(
            "근무표 이미지 변환 시간이 초과되었습니다."
        ) from exc
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or exc.stdout or "").strip()
        raise WorkScheduleConversionError(
            f"근무표 이미지 변환에 실패했습니다. {detail}"
        ) from exc


def _validate_and_make_preview_workbook(source_path, preview_path):
    try:
        workbook = load_workbook(source_path)
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise WorkScheduleValidationError("정상적인 .xlsx 파일이 아닙니다.") from exc

    try:
        if SCHEDULE_SHEET_NAME not in workbook.sheetnames:
            raise WorkScheduleValidationError(
                f"'{SCHEDULE_SHEET_NAME}' 시트가 필요합니다."
            )

        for worksheet in workbook.worksheets:
            worksheet.sheet_state = "visible"
        workbook.save(preview_path)
    finally:
        workbook.close()


def _render_preview_pages(preview_workbook_path, temp_dir):
    libreoffice_path = getattr(settings, "LIBREOFFICE_PATH", "soffice")
    pdftoppm_path = getattr(settings, "PDFTOPPM_PATH", "pdftoppm")
    timeout = getattr(settings, "WORK_SCHEDULE_CONVERSION_TIMEOUT", 180)

    _run_command(
        [
            libreoffice_path,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(temp_dir),
            str(preview_workbook_path),
        ],
        timeout=timeout,
    )

    pdf_path = Path(temp_dir) / f"{Path(preview_workbook_path).stem}.pdf"
    if not pdf_path.exists():
        raise WorkScheduleConversionError("근무표 PDF가 생성되지 않았습니다.")

    output_prefix = Path(temp_dir) / "schedule-page"
    _run_command(
        [
            pdftoppm_path,
            "-png",
            "-r",
            str(getattr(settings, "WORK_SCHEDULE_PREVIEW_DPI", 150)),
            str(pdf_path),
            str(output_prefix),
        ],
        timeout=timeout,
    )

    page_paths = sorted(
        Path(temp_dir).glob("schedule-page-*.png"),
        key=lambda path: int(path.stem.rsplit("-", 1)[1]),
    )
    if not page_paths:
        raise WorkScheduleConversionError(
            "근무표 미리보기 이미지가 생성되지 않았습니다."
        )

    pages = []
    for page_path in page_paths:
        content = page_path.read_bytes()
        width, height = _png_dimensions(content)
        pages.append(PreviewPage(content=content, width=width, height=height))
    return pages


def prepare_work_schedule(uploaded_file):
    original_name = Path(uploaded_file.name or "").name
    if Path(original_name).suffix.lower() != ".xlsx":
        raise WorkScheduleValidationError(".xlsx 파일만 업로드할 수 있습니다.")

    max_size = getattr(settings, "WORK_SCHEDULE_MAX_UPLOAD_SIZE", 20 * 1024 * 1024)
    max_size_mb = max_size // (1024 * 1024)
    size_error = f"파일 크기는 {max_size_mb}MB를 초과할 수 없습니다."
    if uploaded_file.size > max_size:
        raise WorkScheduleValidationError(size_error)

    uploaded_file.seek(0)
    original_content = uploaded_file.read(max_size + 1)
    if len(original_content) > max_size:
        raise WorkScheduleValidationError(size_error)
    if original_content[:4] != b"PK\x03\x04":
        raise WorkScheduleValidationError("정상적인 .xlsx 파일이 아닙니다.")

    with tempfile.TemporaryDirectory(prefix="work-schedule-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        source_path = temp_dir / "source.xlsx"
        preview_path = temp_dir / "preview.xlsx"
        source_path.write_bytes(original_content)
        _validate_and_make_preview_workbook(source_path, preview_path)
        preview_pages = _render_preview_pages(preview_path, temp_dir)

    return PreparedWorkSchedule(
        original_content=original_content,
        original_name=original_name,
        original_size=len(original_content),
        preview_pages=preview_pages,
    )


def delete_storage_files(storage, names):
    for name in filter(None, names):
        try:
            storage.delete(name)
        except OSError:
            pass
